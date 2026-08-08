from __future__ import annotations

import csv
import sys
import unittest
from pathlib import Path

import numpy as np
import openpyxl


ROOT = Path(__file__).resolve().parents[1]
COMMON_DIR = ROOT / "公共代码"
if str(COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_DIR))

from fast_geometry import (  # noqa: E402
    NUMBA_AVAILABLE,
    fast_cylinder_classify,
    fast_cylinder_distance_diagnostics,
    warm_up_fast_geometry,
)
from geometry_kernel import Cylinder, distance_bounds  # noqa: E402


ABSOLUTE_TOLERANCE = 1e-10
RELATIVE_TOLERANCE = 1e-13
MAX_ITERATIONS = 512
CUTOFF = 1.8


def _random_cylinder(rng: np.random.Generator) -> Cylinder:
    return Cylinder(
        center=rng.uniform(-5000.0, 5000.0, 3),
        axis=rng.normal(size=3),
        half_length=float(rng.uniform(1.0, 2500.0)),
        radius=float(rng.uniform(1.0, 50.0)),
    )


def _assert_compatible_bounds(
    test: unittest.TestCase,
    reference,
    candidate,
    tolerance: float = 2e-8,
) -> None:
    test.assertGreaterEqual(candidate.lower, 0.0)
    test.assertGreaterEqual(candidate.upper, candidate.lower)
    test.assertLessEqual(
        max(reference.lower, candidate.lower),
        min(reference.upper, candidate.upper) + tolerance,
    )


@unittest.skipUnless(NUMBA_AVAILABLE, "Numba is required for the accelerated path")
class FastGeometryTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        warm_up_fast_geometry()

    def test_parallel_intersection_and_end_face_cases(self) -> None:
        cases = [
            (
                Cylinder.from_endpoints([0, 0, -2500], [0, 0, 2500], 30),
                Cylinder.from_endpoints([61.8, 0, -2500], [61.8, 0, 2500], 30),
            ),
            (
                Cylinder.from_endpoints([-2500, 0, 0], [2500, 0, 0], 30),
                Cylinder.from_endpoints([0, -2500, 0], [0, 2500, 0], 30),
            ),
            (
                Cylinder.from_endpoints([0, 0, -2500], [0, 0, 2500], 30),
                Cylinder.from_endpoints([0, 0, 2561.8], [0, 0, 7561.8], 30),
            ),
            (
                Cylinder([10, -20, 30], [1, 2, 3], 700, 30),
                Cylinder([900, 300, -500], [-2, 1, 4], 300, 30),
            ),
        ]
        for first, second in cases:
            with self.subTest(first=first.center, second=second.center):
                reference = distance_bounds(
                    first,
                    second,
                    ABSOLUTE_TOLERANCE,
                    RELATIVE_TOLERANCE,
                    MAX_ITERATIONS,
                )
                accelerated = fast_cylinder_distance_diagnostics(
                    first,
                    second,
                    ABSOLUTE_TOLERANCE,
                    RELATIVE_TOLERANCE,
                    MAX_ITERATIONS,
                    cutoff=CUTOFF,
                )
                self.assertEqual(
                    reference.classify(CUTOFF),
                    accelerated.bounds.classify(CUTOFF),
                )
                _assert_compatible_bounds(self, reference, accelerated.bounds)

    def test_threshold_neighbourhood_matches_and_guard_falls_back(self) -> None:
        first = Cylinder.from_endpoints([0, 0, -2500], [0, 0, 2500], 30)
        deltas = (-1e-7, -1e-9, 0.0, 1e-9, 1e-7)
        for delta in deltas:
            second = Cylinder.from_endpoints(
                [60.0 + CUTOFF + delta, 0, -2500],
                [60.0 + CUTOFF + delta, 0, 2500],
                30,
            )
            result = fast_cylinder_distance_diagnostics(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
                cutoff=CUTOFF,
            )
            reference = distance_bounds(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
            )
            self.assertEqual(reference.classify(CUTOFF), result.bounds.classify(CUTOFF))
            if abs(delta) <= 1e-9:
                self.assertTrue(result.used_fallback)
                self.assertEqual(result.fallback_reason, "threshold_guard")

        for delta in deltas:
            second = Cylinder.from_endpoints(
                [0, 0, 2500.0 + CUTOFF + delta],
                [0, 0, 7500.0 + CUTOFF + delta],
                30,
            )
            result = fast_cylinder_classify(
                first,
                second,
                CUTOFF,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
            )
            reference = distance_bounds(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
            )
            self.assertEqual(reference.classify(CUTOFF), result.classification)

    def test_unthresholded_random_distance_bounds_overlap_reference(self) -> None:
        rng = np.random.default_rng(2026080711)
        for pair_index in range(250):
            first = _random_cylinder(rng)
            if pair_index % 2:
                second = _random_cylinder(rng)
            else:
                second = Cylinder(
                    first.center + rng.normal(0.0, 1200.0, 3),
                    rng.normal(size=3),
                    float(rng.uniform(1.0, 2500.0)),
                    float(rng.uniform(1.0, 50.0)),
                )
            reference = distance_bounds(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
            )
            accelerated = fast_cylinder_distance_diagnostics(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
            )
            _assert_compatible_bounds(self, reference, accelerated.bounds)

    def test_large_random_threshold_classification_matches_reference(self) -> None:
        rng = np.random.default_rng(2026080713)
        fallback_count = 0
        for pair_index in range(750):
            first = _random_cylinder(rng)
            if pair_index % 3:
                second = Cylinder(
                    first.center + rng.normal(0.0, 1400.0, 3),
                    rng.normal(size=3),
                    float(rng.uniform(1.0, 2500.0)),
                    float(rng.uniform(1.0, 50.0)),
                )
            else:
                second = _random_cylinder(rng)
            reference = distance_bounds(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
            )
            accelerated = fast_cylinder_distance_diagnostics(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
                cutoff=CUTOFF,
            )
            fallback_count += int(accelerated.used_fallback)
            self.assertEqual(
                reference.classify(CUTOFF),
                accelerated.bounds.classify(CUTOFF),
            )
            _assert_compatible_bounds(self, reference, accelerated.bounds)
        self.assertLess(fallback_count, 20)

    def test_iteration_limit_is_rechecked_by_reference(self) -> None:
        first = Cylinder([0, 0, 0], [1, 2, 3], 2500, 30)
        second = Cylinder([4000, 3000, -2000], [-2, 1, 4], 2500, 30)
        result = fast_cylinder_distance_diagnostics(
            first,
            second,
            ABSOLUTE_TOLERANCE,
            RELATIVE_TOLERANCE,
            1,
        )
        self.assertTrue(result.used_fallback)
        self.assertEqual(result.fallback_reason, "iteration_limit")

    def test_q1_actual_narrow_phase_pairs_match_reference(self) -> None:
        workbook_path = ROOT / "00_赛题与附件" / "附件.xlsx"
        narrow_path = ROOT / "问题" / "问题1" / "results" / "narrow_phase_pairs.csv"
        self.assertTrue(workbook_path.exists())
        self.assertTrue(narrow_path.exists())

        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
        segments: dict[tuple[str, int], Cylinder] = {}
        try:
            for sheet in workbook.sheetnames:
                for row_index, values in enumerate(
                    workbook[sheet].iter_rows(
                        min_row=3,
                        max_col=6,
                        values_only=True,
                    ),
                    start=3,
                ):
                    if all(value is None for value in values):
                        continue
                    segments[(sheet, row_index)] = Cylinder.from_endpoints(
                        values[:3],
                        values[3:],
                        30.0,
                    )
        finally:
            workbook.close()

        with narrow_path.open(encoding="utf-8-sig", newline="") as stream:
            records = list(csv.DictReader(stream))
        self.assertEqual(len(records), 195)
        for record in records:
            first = segments[(record["sheet"], int(record["row_u"]))]
            second = segments[(record["sheet"], int(record["row_v"]))]
            reference = distance_bounds(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
            )
            accelerated = fast_cylinder_distance_diagnostics(
                first,
                second,
                ABSOLUTE_TOLERANCE,
                RELATIVE_TOLERANCE,
                MAX_ITERATIONS,
                cutoff=CUTOFF,
            )
            self.assertEqual(
                reference.classify(CUTOFF),
                accelerated.bounds.classify(CUTOFF),
                msg=f"{record['sheet']} rows {record['row_u']}/{record['row_v']}",
            )
            _assert_compatible_bounds(self, reference, accelerated.bounds)


if __name__ == "__main__":
    unittest.main()
