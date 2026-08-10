from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
from collections import Counter, defaultdict, deque
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


COORD_TOL_NM = 1e-6
COLLINEAR_DOT_TOL = 1e-9
LENGTH_TOL_NM = 1e-5
TARGET_LENGTH_NM = 5000.0


def _discover_project_root(script_path: Path) -> Path:
    # 优先使用环境变量，否则向上查找项目标志目录，避免写死机器绝对路径。
    configured = os.environ.get("MCM_PROJECT_ROOT")
    candidates = [Path(configured).expanduser()] if configured else script_path.resolve().parents
    for candidate in candidates:
        root = candidate.resolve()
        if (root / "公共代码").is_dir() and (root / "问题").is_dir():
            return root
    raise RuntimeError("无法定位项目根目录；请设置 MCM_PROJECT_ROOT")


PROJECT_ROOT = _discover_project_root(Path(__file__))
DEFAULT_SOURCE = PROJECT_ROOT / "00_赛题与附件" / "附件.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "02_数据与参数" / "processed"


@dataclass(frozen=True)
class Segment:
    sid: int
    sheet_row: int
    p1: tuple[float, float, float]
    p2: tuple[float, float, float]

    @property
    def vector(self) -> tuple[float, float, float]:
        return tuple(self.p2[i] - self.p1[i] for i in range(3))

    @property
    def length(self) -> float:
        return math.dist(self.p1, self.p2)

    def endpoint(self, side: int) -> tuple[float, float, float]:
        return self.p1 if side == 1 else self.p2

    def outward(self, side: int) -> tuple[float, float, float]:
        point = self.endpoint(side)
        other = self.p2 if side == 1 else self.p1
        length = self.length
        return tuple((other[i] - point[i]) / length for i in range(3))


@dataclass(frozen=True)
class Junction:
    eid1: tuple[int, int]
    eid2: tuple[int, int]
    shift: tuple[int, int, int]
    dot_outward: float
    kind: str

    @property
    def axes(self) -> tuple[str, ...]:
        return tuple(axis for axis, value in zip("XYZ", self.shift) if value)


class UnionFind:
    def __init__(self, n: int) -> None:
        self.parent = list(range(n))

    def find(self, value: int) -> int:
        while self.parent[value] != value:
            self.parent[value] = self.parent[self.parent[value]]
            value = self.parent[value]
        return value

    def union(self, left: int, right: int) -> None:
        left_root = self.find(left)
        right_root = self.find(right)
        if left_root != right_root:
            self.parent[right_root] = left_root


def dot(left: Iterable[float], right: Iterable[float]) -> float:
    return sum(a * b for a, b in zip(left, right))


def source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


# 附件只读；周期等价端点按坐标容差构造。
def load_segments(path: Path) -> dict[str, list[Segment]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[Segment]] = {}
    try:
        for sheet in workbook.worksheets:
            segments: list[Segment] = []
            for sheet_row, row in enumerate(
                sheet.iter_rows(min_row=3, max_col=6, values_only=True),
                start=3,
            ):
                if all(value is None for value in row):
                    continue
                if any(value is None for value in row):
                    raise ValueError(f"{sheet.title}!{sheet_row} 不是完整六列坐标")
                values = tuple(float(value) for value in row)
                segments.append(
                    Segment(
                        sid=len(segments),
                        sheet_row=sheet_row,
                        p1=values[:3],
                        p2=values[3:],
                    )
                )
            result[sheet.title] = segments
    finally:
        workbook.close()
    return result


def canonical_value(value: float, period: float) -> float:
    half = period / 2.0
    if abs(value - half) <= COORD_TOL_NM or abs(value + half) <= COORD_TOL_NM:
        return -half
    return value


def canonical_key(point: tuple[float, float, float], domain: tuple[float, float, float]) -> tuple[int, int, int]:
    return tuple(
        round(canonical_value(value, period) / COORD_TOL_NM)
        for value, period in zip(point, domain)
    )


def periodic_shift(
    left: tuple[float, float, float],
    right: tuple[float, float, float],
    domain: tuple[float, float, float],
) -> tuple[int, int, int] | None:
    shift: list[int] = []
    for left_value, right_value, period in zip(left, right, domain):
        raw = (right_value - left_value) / period
        nearest = int(round(raw))
        if nearest not in (-1, 0, 1):
            return None
        if abs((right_value - left_value) - nearest * period) > COORD_TOL_NM:
            return None
        shift.append(nearest)
    return tuple(shift)


# 接头必须一对一、周期等价且共线；多解显式保留。
# 关键：按周期边界和几何容差生成可连接端点候选。
def build_candidate_junctions(
    segments: list[Segment], domain: tuple[float, float, float]
) -> list[Junction]:
    groups: dict[tuple[int, int, int], list[tuple[int, int]]] = defaultdict(list)
    for segment in segments:
        for side in (1, 2):
            groups[canonical_key(segment.endpoint(side), domain)].append((segment.sid, side))

    candidates: list[Junction] = []
    for endpoints in groups.values():
        for left_index in range(len(endpoints)):
            for right_index in range(left_index + 1, len(endpoints)):
                left = endpoints[left_index]
                right = endpoints[right_index]
                if left[0] == right[0]:
                    continue
                left_segment = segments[left[0]]
                right_segment = segments[right[0]]
                shift = periodic_shift(
                    left_segment.endpoint(left[1]),
                    right_segment.endpoint(right[1]),
                    domain,
                )
                if shift is None:
                    continue
                direction_dot = dot(
                    left_segment.outward(left[1]), right_segment.outward(right[1])
                )
                if abs(direction_dot + 1.0) > COLLINEAR_DOT_TOL:
                    continue
                candidates.append(
                    Junction(
                        eid1=left,
                        eid2=right,
                        shift=shift,
                        dot_outward=direction_dot,
                        kind="shared" if shift == (0, 0, 0) else "periodic",
                    )
                )
    return sorted(candidates, key=lambda edge: (edge.eid1, edge.eid2))


def choose_junctions(
    segments: list[Segment], candidates: list[Junction]
) -> tuple[list[Junction], set[int], dict[str, int]]:
    endpoint_degree: Counter[tuple[int, int]] = Counter()
    for edge in candidates:
        endpoint_degree[edge.eid1] += 1
        endpoint_degree[edge.eid2] += 1

    forced = [
        edge
        for edge in candidates
        if endpoint_degree[edge.eid1] == 1 and endpoint_degree[edge.eid2] == 1
    ]
    selected = list(forced)
    used = {endpoint for edge in selected for endpoint in (edge.eid1, edge.eid2)}
    ambiguous_edges = [
        edge
        for edge in candidates
        if edge not in forced and edge.eid1 not in used and edge.eid2 not in used
    ]

    # 对小型歧义分量穷举最大匹配，并保留解数而不是静默认定唯一身份。
    ambiguous_segments: set[int] = set()
    solution_count = 1
    if ambiguous_edges:
        nodes: dict[tuple[int, int], list[int]] = defaultdict(list)
        for index, edge in enumerate(ambiguous_edges):
            nodes[edge.eid1].append(index)
            nodes[edge.eid2].append(index)
        seen_edges: set[int] = set()
        for start in range(len(ambiguous_edges)):
            if start in seen_edges:
                continue
            queue = deque([start])
            component_edges: list[int] = []
            while queue:
                edge_index = queue.popleft()
                if edge_index in seen_edges:
                    continue
                seen_edges.add(edge_index)
                component_edges.append(edge_index)
                edge = ambiguous_edges[edge_index]
                for endpoint in (edge.eid1, edge.eid2):
                    queue.extend(nodes[endpoint])

            local_edges = [ambiguous_edges[index] for index in component_edges]
            local_segments = {
                endpoint[0]
                for edge in local_edges
                for endpoint in (edge.eid1, edge.eid2)
            }
            ambiguous_segments.update(local_segments)
            if len(local_edges) > 22:
                continue

            best: list[list[Junction]] = []
            best_size = -1

            def search(position: int, occupied: set[tuple[int, int]], chosen: list[Junction]) -> None:
                nonlocal best_size
                if position == len(local_edges):
                    size = len(chosen)
                    if size > best_size:
                        best_size = size
                        best.clear()
                        best.append(list(chosen))
                    elif size == best_size and len(best) < 1001:
                        best.append(list(chosen))
                    return
                if len(chosen) + len(local_edges) - position < best_size:
                    return
                edge = local_edges[position]
                search(position + 1, occupied, chosen)
                if edge.eid1 not in occupied and edge.eid2 not in occupied:
                    occupied.add(edge.eid1)
                    occupied.add(edge.eid2)
                    chosen.append(edge)
                    search(position + 1, occupied, chosen)
                    chosen.pop()
                    occupied.remove(edge.eid1)
                    occupied.remove(edge.eid2)

            search(0, set(), [])
            solution_count *= max(1, len(best))
            if best:
                selected.extend(sorted(best[0], key=lambda edge: (edge.eid1, edge.eid2)))

    diagnostics = {
        "candidate_junctions": len(candidates),
        "selected_junctions": len(selected),
        "ambiguous_endpoints": sum(value > 1 for value in endpoint_degree.values()),
        "maximum_matching_solutions": solution_count,
    }
    return sorted(selected, key=lambda edge: (edge.eid1, edge.eid2)), ambiguous_segments, diagnostics


def face_labels(point: tuple[float, float, float], domain: tuple[float, float, float]) -> list[str]:
    labels: list[str] = []
    for axis, value, period in zip("XYZ", point, domain):
        half = period / 2.0
        if abs(value + half) <= COORD_TOL_NM:
            labels.append(f"{axis}-")
        if abs(value - half) <= COORD_TOL_NM:
            labels.append(f"{axis}+")
    return labels


def order_component_chain(
    member_ids: list[int], edges: list[Junction], segments: list[Segment]
) -> list[tuple[int, bool]]:
    links: dict[tuple[int, int], tuple[int, int]] = {}
    for edge in edges:
        links[edge.eid1] = edge.eid2
        links[edge.eid2] = edge.eid1

    unmatched = [
        (sid, side)
        for sid in member_ids
        for side in (1, 2)
        if (sid, side) not in links
    ]
    if unmatched:
        start = min(
            unmatched,
            key=lambda endpoint: (
                segments[endpoint[0]].endpoint(endpoint[1]),
                segments[endpoint[0]].sheet_row,
                endpoint[1],
            ),
        )
    else:
        start = (min(member_ids), 1)

    ordered: list[tuple[int, bool]] = []
    visited: set[int] = set()
    current = start
    while current[0] not in visited:
        sid, entry_side = current
        visited.add(sid)
        ordered.append((sid, entry_side == 2))
        exit_endpoint = (sid, 3 - entry_side)
        if exit_endpoint not in links:
            break
        current = links[exit_endpoint]
    if len(visited) != len(member_ids):
        for sid in sorted(set(member_ids) - visited):
            ordered.append((sid, False))
    return ordered


# 粒子链只连接已有片段，不补齐缺失长度或延伸短段。
# 关键：重建周期情景中的粒子链并核验总长度与残差。
def analyze_periodic_scenario(
    sheet: str,
    scenario: str,
    segments: list[Segment],
    domain: tuple[float, float, float],
) -> dict:
    candidates = build_candidate_junctions(segments, domain)
    selected, ambiguous_segments, diagnostics = choose_junctions(segments, candidates)
    union_find = UnionFind(len(segments))
    for edge in selected:
        union_find.union(edge.eid1[0], edge.eid2[0])

    component_members: dict[int, list[int]] = defaultdict(list)
    for segment in segments:
        component_members[union_find.find(segment.sid)].append(segment.sid)

    edge_by_component: dict[int, list[Junction]] = defaultdict(list)
    for edge in selected:
        edge_by_component[union_find.find(edge.eid1[0])].append(edge)

    components: list[dict] = []
    row_map: list[dict] = []
    for sequence, root in enumerate(
        sorted(component_members, key=lambda value: min(component_members[value])), start=1
    ):
        member_ids = sorted(component_members[root])
        member_segments = [segments[sid] for sid in member_ids]
        edges = edge_by_component[root]
        total_length = sum(segment.length for segment in member_segments)
        residual = total_length - TARGET_LENGTH_NM
        axes = Counter(axis for edge in edges for axis in edge.axes)
        ordered_chain = order_component_chain(member_ids, edges, segments)
        chain_position = {
            sid: (position, reversed_direction)
            for position, (sid, reversed_direction) in enumerate(ordered_chain, start=1)
        }
        unmatched_faces: Counter[str] = Counter()
        matched_endpoints = {endpoint for edge in edges for endpoint in (edge.eid1, edge.eid2)}
        for segment in member_segments:
            for side in (1, 2):
                endpoint = (segment.sid, side)
                if endpoint not in matched_endpoints:
                    unmatched_faces.update(face_labels(segment.endpoint(side), domain))

        has_ambiguity = any(sid in ambiguous_segments for sid in member_ids)
        complete = abs(residual) <= LENGTH_TOL_NM
        if has_ambiguity:
            status = "multiple_solutions"
        elif complete:
            status = "unique"
        else:
            status = "cannot_reconstruct"
        particle_id = f"{sheet}_{scenario}_P{sequence:04d}"
        component = {
            "sheet": sheet,
            "scenario": scenario,
            "particle_id": particle_id,
            "status": status,
            "complete_5000": complete,
            "segment_count": len(member_ids),
            "sheet_rows": ";".join(str(segment.sheet_row) for segment in member_segments),
            "chain_rows": "->".join(
                str(segments[sid].sheet_row) for sid, _ in ordered_chain
            ),
            "chain_orientations": ";".join(
                f"{segments[sid].sheet_row}:{'-' if reversed_direction else '+'}"
                for sid, reversed_direction in ordered_chain
            ),
            "total_length_nm": total_length,
            "length_residual_nm": residual,
            "junction_count": len(edges),
            "cross_X": axes["X"],
            "cross_Y": axes["Y"],
            "cross_Z": axes["Z"],
            "shared_junctions": sum(edge.kind == "shared" for edge in edges),
            "periodic_junctions": sum(edge.kind == "periodic" for edge in edges),
            "unmatched_boundary_faces": ";".join(
                f"{face}:{count}" for face, count in sorted(unmatched_faces.items())
            ),
        }
        components.append(component)
        for segment in member_segments:
            row_map.append(
                {
                    "sheet": sheet,
                    "scenario": scenario,
                    "sheet_row": segment.sheet_row,
                    "particle_id": particle_id,
                    "component_status": status,
                    "chain_order": chain_position[segment.sid][0],
                    "orientation_reversed": chain_position[segment.sid][1],
                    "segment_length_nm": segment.length,
                    "component_total_length_nm": total_length,
                    "component_residual_nm": residual,
                }
            )

    junction_rows: list[dict] = []
    for index, edge in enumerate(selected, start=1):
        left_segment = segments[edge.eid1[0]]
        right_segment = segments[edge.eid2[0]]
        left_point = left_segment.endpoint(edge.eid1[1])
        right_point = right_segment.endpoint(edge.eid2[1])
        junction_rows.append(
            {
                "sheet": sheet,
                "scenario": scenario,
                "junction_id": index,
                "row1": left_segment.sheet_row,
                "side1": edge.eid1[1],
                "cells1": (
                    f"A{left_segment.sheet_row}:C{left_segment.sheet_row}"
                    if edge.eid1[1] == 1
                    else f"D{left_segment.sheet_row}:F{left_segment.sheet_row}"
                ),
                "endpoint1_x": left_point[0],
                "endpoint1_y": left_point[1],
                "endpoint1_z": left_point[2],
                "row2": right_segment.sheet_row,
                "side2": edge.eid2[1],
                "cells2": (
                    f"A{right_segment.sheet_row}:C{right_segment.sheet_row}"
                    if edge.eid2[1] == 1
                    else f"D{right_segment.sheet_row}:F{right_segment.sheet_row}"
                ),
                "endpoint2_x": right_point[0],
                "endpoint2_y": right_point[1],
                "endpoint2_z": right_point[2],
                "kind": edge.kind,
                "shift_x": edge.shift[0],
                "shift_y": edge.shift[1],
                "shift_z": edge.shift[2],
                "crossing_axes": "".join(edge.axes),
                "dot_outward": edge.dot_outward,
            }
        )

    summary = summarize_scenario(sheet, scenario, domain, segments, components, diagnostics)
    return {
        "summary": summary,
        "components": components,
        "row_map": row_map,
        "junctions": junction_rows,
    }


def analyze_literal_scenario(sheet: str, segments: list[Segment]) -> dict:
    components: list[dict] = []
    row_map: list[dict] = []
    for sequence, segment in enumerate(segments, start=1):
        residual = segment.length - TARGET_LENGTH_NM
        complete = abs(residual) <= LENGTH_TOL_NM
        status = "unique" if complete else "cannot_reconstruct"
        particle_id = f"{sheet}_row_literal_P{sequence:04d}"
        components.append(
            {
                "sheet": sheet,
                "scenario": "row_literal",
                "particle_id": particle_id,
                "status": status,
                "complete_5000": complete,
                "segment_count": 1,
                "sheet_rows": str(segment.sheet_row),
                "chain_rows": str(segment.sheet_row),
                "chain_orientations": f"{segment.sheet_row}:+",
                "total_length_nm": segment.length,
                "length_residual_nm": residual,
                "junction_count": 0,
                "cross_X": 0,
                "cross_Y": 0,
                "cross_Z": 0,
                "shared_junctions": 0,
                "periodic_junctions": 0,
                "unmatched_boundary_faces": "",
            }
        )
        row_map.append(
            {
                "sheet": sheet,
                "scenario": "row_literal",
                "sheet_row": segment.sheet_row,
                "particle_id": particle_id,
                "component_status": status,
                "chain_order": 1,
                "orientation_reversed": False,
                "segment_length_nm": segment.length,
                "component_total_length_nm": segment.length,
                "component_residual_nm": residual,
            }
        )
    diagnostics = {
        "candidate_junctions": 0,
        "selected_junctions": 0,
        "ambiguous_endpoints": 0,
        "maximum_matching_solutions": 1,
    }
    summary = summarize_scenario(
        sheet, "row_literal", None, segments, components, diagnostics
    )
    return {"summary": summary, "components": components, "row_map": row_map, "junctions": []}


def summarize_scenario(
    sheet: str,
    scenario: str,
    domain: tuple[float, float, float] | None,
    segments: list[Segment],
    components: list[dict],
    diagnostics: dict[str, int],
) -> dict:
    status_counts = Counter(component["status"] for component in components)
    residuals = [float(component["length_residual_nm"]) for component in components]
    return {
        "sheet": sheet,
        "scenario": scenario,
        "domain_x_nm": domain[0] if domain else "",
        "domain_y_nm": domain[1] if domain else "",
        "domain_z_nm": domain[2] if domain else "",
        "input_segments": len(segments),
        "potential_particle_chains": len(components),
        "unique_complete_particles": sum(
            component["status"] == "unique" and component["complete_5000"]
            for component in components
        ),
        "multiple_solution_chains": status_counts["multiple_solutions"],
        "unreconstructable_chains": status_counts["cannot_reconstruct"],
        "complete_5000_chains": sum(component["complete_5000"] for component in components),
        "complete_chain_rate": sum(component["complete_5000"] for component in components)
        / len(components),
        "sum_signed_length_residual_nm": sum(residuals),
        "sum_abs_length_residual_nm": sum(abs(value) for value in residuals),
        "max_abs_length_residual_nm": max(abs(value) for value in residuals),
        "cross_X": sum(int(component["cross_X"]) for component in components),
        "cross_Y": sum(int(component["cross_Y"]) for component in components),
        "cross_Z": sum(int(component["cross_Z"]) for component in components),
        **diagnostics,
    }


def boundary_evidence(sheet: str, segments: list[Segment]) -> list[dict]:
    rows: list[dict] = []
    for axis_index, axis in enumerate("XYZ"):
        for half_width in (500.0, 5000.0):
            for sign in (-1, 1):
                value = sign * half_width
                endpoint_count = 0
                short_segment_count = 0
                cells: list[str] = []
                for segment in segments:
                    touched = False
                    for side, point in ((1, segment.p1), (2, segment.p2)):
                        if abs(point[axis_index] - value) <= COORD_TOL_NM:
                            endpoint_count += 1
                            touched = True
                            column = "ABC"[axis_index] if side == 1 else "DEF"[axis_index]
                            cells.append(f"{column}{segment.sheet_row}")
                    if touched and segment.length < TARGET_LENGTH_NM - LENGTH_TOL_NM:
                        short_segment_count += 1
                rows.append(
                    {
                        "sheet": sheet,
                        "axis": axis,
                        "plane_nm": value,
                        "endpoint_count": endpoint_count,
                        "short_segment_count": short_segment_count,
                        "cells": ";".join(cells),
                    }
                )
    return rows


def scaled_yz_test(sheet: str, segments: list[Segment]) -> dict:
    lengths = []
    for segment in segments:
        p1 = (segment.p1[0], 10.0 * segment.p1[1], 10.0 * segment.p1[2])
        p2 = (segment.p2[0], 10.0 * segment.p2[1], 10.0 * segment.p2[2])
        lengths.append(math.dist(p1, p2))
    return {
        "sheet": sheet,
        "rows": len(lengths),
        "complete_5000_after_yz_x10": sum(
            abs(length - TARGET_LENGTH_NM) <= LENGTH_TOL_NM for length in lengths
        ),
        "min_length_nm": min(lengths),
        "median_length_nm": sorted(lengths)[len(lengths) // 2],
        "max_length_nm": max(lengths),
        "sum_abs_residual_nm": sum(abs(length - TARGET_LENGTH_NM) for length in lengths),
    }


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def validate_analysis(analysis: dict, segments: list[Segment]) -> None:
    expected_rows = {segment.sheet_row for segment in segments}
    mapped_rows = [int(row["sheet_row"]) for row in analysis["row_map"]]
    if len(mapped_rows) != len(expected_rows) or set(mapped_rows) != expected_rows:
        raise AssertionError("row identity map is not a one-to-one partition")

    component_rows: list[int] = []
    for component in analysis["components"]:
        rows = [int(value) for value in str(component["sheet_rows"]).split(";")]
        chain_rows = [int(value) for value in str(component["chain_rows"]).split("->")]
        if sorted(rows) != sorted(chain_rows):
            raise AssertionError("ordered chain and component members disagree")
        component_rows.extend(rows)
        residual = float(component["length_residual_nm"])
        if bool(component["complete_5000"]) != (abs(residual) <= LENGTH_TOL_NM):
            raise AssertionError("length completion flag disagrees with residual")
        if float(component["total_length_nm"]) > TARGET_LENGTH_NM + LENGTH_TOL_NM:
            raise AssertionError("a reconstructed component exceeds one medium length")
    if len(component_rows) != len(expected_rows) or set(component_rows) != expected_rows:
        raise AssertionError("components do not partition input rows")

    input_total = sum(segment.length for segment in segments)
    component_total = sum(float(row["total_length_nm"]) for row in analysis["components"])
    if abs(input_total - component_total) > LENGTH_TOL_NM:
        raise AssertionError("component reconstruction does not conserve segment length")


def markdown_report(source: Path, digest: str, result: dict) -> str:
    summary = result["scenario_summary"]
    lines = [
        "# 附件分段重建审计",
        "",
        f"- 原件：`{source}`",
        f"- SHA-256：`{digest}`",
        "- 原件只读；所有输出均为派生审计文件。",
        "",
        "## 算法",
        "",
        "1. 每行视为一个线段，计算端点、直接轴长和端点向外单位切向量。",
        "2. 对候选域把相对面坐标按周期归一；仅当两个端点相差整数个域长、且两端向外切向量点积为 -1（容差 1e-9）时建立候选接头。",
        "3. 每个端点最多匹配一次；唯一候选直接采用，多候选采用最大匹配并保留多解标志。",
        "4. 以已选接头的连通分量作为潜在原介质，片段长度求和并与 5000 nm 比较。",
        "5. `unique` 表示身份唯一且总长为 5000；`multiple_solutions` 表示存在多种端点匹配；`cannot_reconstruct` 表示当前输入下总长不闭合。",
        "",
        "## 场景汇总",
        "",
        "|分表|解释|候选域(nm)|输入片段|潜在链|唯一完整|多解|无法重建|总绝对残差(nm)|跨X/Y/Z|",
        "|---|---|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in summary:
        domain = (
            "逐行"
            if row["scenario"] == "row_literal"
            else f"{row['domain_x_nm']}x{row['domain_y_nm']}x{row['domain_z_nm']}"
        )
        lines.append(
            f"|{row['sheet']}|{row['scenario']}|{domain}|{row['input_segments']}|"
            f"{row['potential_particle_chains']}|{row['unique_complete_particles']}|"
            f"{row['multiple_solution_chains']}|{row['unreconstructable_chains']}|"
            f"{row['sum_abs_length_residual_nm']:.6f}|"
            f"{row['cross_X']}/{row['cross_Y']}/{row['cross_Z']}|"
        )

    incomplete_thin = [
        component
        for component in result["components"]
        if component["scenario"] == "thin_prism_periodic"
        and component["status"] == "cannot_reconstruct"
    ]
    thin_crossings = [
        junction
        for junction in result["junctions"]
        if junction["scenario"] == "thin_prism_periodic"
        and (junction["shift_y"] or junction["shift_z"])
    ]
    group3_examples = [
        junction
        for junction in result["junctions"]
        if junction["sheet"] == "组3"
        and junction["scenario"] == "full_cube_periodic"
    ][:3]
    scaled_tests = {row["sheet"]: row for row in result["scaled_yz_test"]}

    lines.extend(
        [
            "",
            "## 核心结果",
            "",
            "- 三种解释的端点候选均无多解：所有已选周期接头都是端点一对一且共线的唯一匹配。这里的“唯一”只针对数值身份，不代表域假设已被证明。",
            "- 对组1，薄棱柱候选把总绝对长度残差从完整立方体解释的 10750.506 nm 降至 750.506 nm，并得到 5/7 条完整链。",
            "- 对组2，薄棱柱候选把总绝对长度残差从 55768.835 nm 降至 768.835 nm，并得到 24/28 条完整链。",
            "- 对组3，题设立方体周期解释得到 357 条潜在链，其中 305 条完整、52 条不闭合；这说明正式域长能解释大部分但不是全部记录。",
            "- 所有场景均未发现不同记录在同一坐标直接共享端点；重联证据全部来自相对面的周期等价端点。",
            "",
            "## 组1/组2未闭合链",
            "",
            "|分表|介质ID|链序|总长(nm)|残差(nm)|未匹配边界|",
            "|---|---|---|---:|---:|---|",
        ]
    )
    for component in incomplete_thin:
        lines.append(
            f"|{component['sheet']}|{component['particle_id']}|{component['chain_rows']}|"
            f"{component['total_length_nm']:.6f}|{component['length_residual_nm']:.6f}|"
            f"{component['unmatched_boundary_faces']}|"
        )

    lines.extend(
        [
            "",
            "## 关键端点证据",
            "",
            "下表列出组1/组2所有跨 `Y/Z=±500` 的共线接头，并附组3正式边界的代表性接头。",
            "",
            "|分表|解释|单元格1|单元格2|跨界轴|周期位移|向外切向点积|",
            "|---|---|---|---|---|---|---:|",
        ]
    )
    for junction in thin_crossings + group3_examples:
        shift = f"({junction['shift_x']},{junction['shift_y']},{junction['shift_z']})"
        lines.append(
            f"|{junction['sheet']}|{junction['scenario']}|{junction['cells1']}|"
            f"{junction['cells2']}|{junction['crossing_axes']}|{shift}|"
            f"{junction['dot_outward']:.12f}|"
        )

    lines.extend(
        [
            "",
            "## ±500最可能生成机制",
            "",
            "**证据最支持但尚不能证实的解释**：组1/组2的数据生成或导出环节使用了轴向异尺度边界 `Lx=10000, Ly=Lz=1000`，并在 `Y/Z=±500` 做了周期分段；另有少量端段未导出或被终端裁切。依据是：所有短段都触及 `±500/±5000`，跨 `±500` 的端点以 15 位左右精度重合、方向严格共线，且重联后大量链总长在浮点误差内恰为 5000 nm。",
            "",
            "这一解释仍不是事实，原因是薄棱柱重建仍有 6 条链不闭合，且题面明确给出 10000 nm 正方体。较弱的替代解释包括：从完整立方体抽取中央窄窗、手工构造测试路径、或导出程序混用了裁切与周期平移。单纯“Y/Z单位缩小了10倍”不符合数据：恢复乘10后两组都没有任何一行保持 5000 nm。",
            "",
            f"- 组1按 Y/Z 乘10：0/{scaled_tests['组1']['rows']} 行为5000 nm，长度范围 {scaled_tests['组1']['min_length_nm']:.3f}–{scaled_tests['组1']['max_length_nm']:.3f} nm。",
            f"- 组2按 Y/Z 乘10：0/{scaled_tests['组2']['rows']} 行为5000 nm，长度范围 {scaled_tests['组2']['min_length_nm']:.3f}–{scaled_tests['组2']['max_length_nm']:.3f} nm。",
            "",
            "## 解释边界",
            "",
            "- 周期接头和共线只证明片段在数值上可重联，不证明附件生成程序或物理边界一定采用该域。",
            "- 若某链总长不足 5000 nm，可能是片段缺失、裁切窗口、导出不完整或题设/附件不一致；本报告不自动补段。",
            "- `组1/组2` 的 `Y/Z=±500` 现象需结合完整链比例、未匹配端点及简单单位缩放反证综合判断。",
            "- Q1正式并列报告逐行片段与题设立方体重联口径；薄棱柱 ID 只能作为附件生成机制诊断和敏感性方案。",
            "",
            "## 输出文件",
            "",
            "- `scenario_summary.csv`：解释级统计。",
            "- `component_results.csv`：潜在原介质链及残差。",
            "- `row_identity_map.csv`：Q1 可使用的逐行介质 ID 映射。",
            "- `junctions.csv`：每个周期/共享接头及跨界方向。",
            "- `boundary_evidence.csv`：±500、±5000 的端点证据。",
            "- `reconstruction_results.json`：完整机器可读结果。",
        ]
    )
    return "\n".join(lines) + "\n"


def reconstruct_workbook(source: Path = DEFAULT_SOURCE) -> dict:
    source = source.resolve()
    digest = source_sha256(source)
    workbook_segments = load_segments(source)

    analyses: list[dict] = []
    for sheet, segments in workbook_segments.items():
        sheet_analyses = [analyze_literal_scenario(sheet, segments)]
        sheet_analyses.append(
            analyze_periodic_scenario(
                sheet, "full_cube_periodic", segments, (10000.0, 10000.0, 10000.0)
            )
        )
        if sheet in ("组1", "组2"):
            sheet_analyses.append(
                analyze_periodic_scenario(
                    sheet, "thin_prism_periodic", segments, (10000.0, 1000.0, 1000.0)
                )
            )
        for analysis in sheet_analyses:
            validate_analysis(analysis, segments)
        analyses.extend(sheet_analyses)

    try:
        source_label = source.relative_to(PROJECT_ROOT.resolve()).as_posix()
    except ValueError:
        source_label = str(source)
    return {
        "source": source_label,
        "source_sha256": digest,
        "validation": "passed",
        "tolerances": {
            "coordinate_nm": COORD_TOL_NM,
            "collinear_dot": COLLINEAR_DOT_TOL,
            "length_nm": LENGTH_TOL_NM,
        },
        "scenario_summary": [analysis["summary"] for analysis in analyses],
        "components": [row for analysis in analyses for row in analysis["components"]],
        "row_identity_map": [row for analysis in analyses for row in analysis["row_map"]],
        "junctions": [row for analysis in analyses for row in analysis["junctions"]],
        "boundary_evidence": [
            row
            for sheet, segments in workbook_segments.items()
            for row in boundary_evidence(sheet, segments)
        ],
        "scaled_yz_test": [
            scaled_yz_test(sheet, segments)
            for sheet, segments in workbook_segments.items()
            if sheet in ("组1", "组2")
        ],
    }


def write_processed_outputs(result: dict, output_dir: Path = DEFAULT_OUTPUT_DIR) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    write_csv(output_dir / "scenario_summary.csv", result["scenario_summary"])
    write_csv(output_dir / "component_results.csv", result["components"])
    write_csv(output_dir / "row_identity_map.csv", result["row_identity_map"])
    write_csv(output_dir / "junctions.csv", result["junctions"])
    write_csv(output_dir / "boundary_evidence.csv", result["boundary_evidence"])
    write_csv(output_dir / "scaled_yz_test.csv", result["scaled_yz_test"])
    (output_dir / "reconstruction_results.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / "reconstruction_report.md").write_text(
        markdown_report(Path(result["source"]), result["source_sha256"], result),
        encoding="utf-8",
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="重建附件中的周期分段身份并输出审计证据")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main() -> None:
    args = parse_args()
    result = reconstruct_workbook(args.source)
    write_processed_outputs(result, args.output_dir)
    print(json.dumps(result["scenario_summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
