from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = PROJECT_ROOT / "00_赛题与附件" / "附件.xlsx"
RESULTS_DIR = PROJECT_ROOT / "问题" / "问题1" / "results"
OUTPUT_DIR = PROJECT_ROOT / "论文" / "figures" / "data"
MAIN_SCENARIO = "A_row_literal"
MAIN_MODE = "disconnected_fragments"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def read_witness_rows() -> dict[str, set[int]]:
    result = {"组1": set(), "组2": set(), "组3": set()}
    path = RESULTS_DIR / "witness_path_edges.csv"
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        for row in csv.DictReader(stream):
            if row["scenario"] != MAIN_SCENARIO or row["internal_mode"] != MAIN_MODE:
                continue
            for field in ("row_u", "row_v"):
                if row[field].strip():
                    result[row["sheet"]].add(int(row[field]))
    return result


def read_conclusions() -> dict[str, bool]:
    result: dict[str, bool] = {}
    path = RESULTS_DIR / "scenario_summary.csv"
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        for row in csv.DictReader(stream):
            if row["scenario"] == MAIN_SCENARIO and row["internal_mode"] == MAIN_MODE:
                result[row["sheet"]] = row["conductive_definite"].strip().lower() == "true"
    if set(result) != {"组1", "组2", "组3"}:
        raise ValueError("问题1主场景结论不完整")
    return result


def workbook_cylinders() -> dict[str, list[dict]]:
    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    result: dict[str, list[dict]] = {}
    try:
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            records: list[dict] = []
            for sheet_row, values in enumerate(
                worksheet.iter_rows(min_row=3, max_col=6, values_only=True), start=3
            ):
                if all(value is None for value in values):
                    continue
                if any(value is None for value in values):
                    raise ValueError(f"{sheet}!{sheet_row} 坐标不完整")
                coordinates = [float(value) for value in values]
                records.append(
                    {
                        "id": f"r{sheet_row}",
                        "start_nm": coordinates[:3],
                        "end_nm": coordinates[3:],
                        "radius_nm": 30.0,
                    }
                )
            result[sheet] = records
    finally:
        workbook.close()
    return result


def build_scenes(output_dir: Path = OUTPUT_DIR) -> list[dict]:
    witnesses = read_witness_rows()
    conclusions = read_conclusions()
    groups = workbook_cylinders()
    output_dir.mkdir(parents=True, exist_ok=True)
    index: list[dict] = []

    for group_index, sheet in enumerate(("组1", "组2", "组3"), start=1):
        cylinders = groups[sheet]
        for cylinder in cylinders:
            row = int(cylinder["id"][1:])
            cylinder["role"] = "witness" if row in witnesses[sheet] else "background_a"

        scene = {
            "name": f"Q1_{sheet}_{MAIN_SCENARIO}_{MAIN_MODE}",
            "publication_status": "formal_result",
            "box": {"length_nm": 10000.0, "show": True, "transparency": 94},
            "electrodes": {"show": True, "thickness_nm": 20.0},
            "cylinders": cylinders,
            "spheres": [],
            "traceability": {
                "source_workbook": "00_赛题与附件/附件.xlsx",
                "source_workbook_sha256": sha256(WORKBOOK),
                "witness_source": "问题/问题1/results/witness_path_edges.csv",
                "witness_source_sha256": sha256(RESULTS_DIR / "witness_path_edges.csv"),
                "scenario": MAIN_SCENARIO,
                "internal_mode": MAIN_MODE,
                "sheet": sheet,
                "conductive": conclusions[sheet],
                "witness_rows": sorted(witnesses[sheet]),
            },
        }
        scene_path = output_dir / f"q1_group{group_index}_scene.json"
        scene_path.write_text(
            json.dumps(scene, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        index.append(
            {
                "sheet": sheet,
                "scene": str(scene_path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                "row_count": len(cylinders),
                "witness_rows": sorted(witnesses[sheet]),
                "conductive": conclusions[sheet],
            }
        )

    (output_dir / "q1_scene_index.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return index


if __name__ == "__main__":
    print(json.dumps(build_scenes(), ensure_ascii=False, indent=2))
