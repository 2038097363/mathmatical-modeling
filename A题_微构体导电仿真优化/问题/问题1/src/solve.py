from __future__ import annotations

import argparse
import csv
import hashlib
import heapq
import json
import platform
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import openpyxl
from scipy.optimize import minimize


PROJECT_ROOT = Path(__file__).resolve().parents[3]
QUESTION_ROOT = Path(__file__).resolve().parents[1]
COMMON_DIR = PROJECT_ROOT / "公共代码"
RECONSTRUCTION_DIR = PROJECT_ROOT / "02_数据与参数" / "src"
for module_dir in (COMMON_DIR, RECONSTRUCTION_DIR):
    if str(module_dir) not in sys.path:
        sys.path.insert(0, str(module_dir))

from geometry_kernel import (
    Cylinder,
    capsule_cylinder_distance,
    distance_bounds,
    shape_plane_distance,
)
from reconstruct_segments import reconstruct_workbook, write_processed_outputs
from result_registry import (
    DEFAULT_REGISTRY,
    DEFAULT_TEX,
    export_latex,
    read_registry,
    register_result,
    write_registry,
)


RADIUS_NM = 30.0
CONTACT_CUTOFF_NM = 1.8
ELECTRODE_X_NM = 5000.0
GJK_ABS_TOL_NM = 1e-10
GJK_REL_TOL = 1e-13
SCREEN_GUARD_NM = 1e-8

SCENARIOS = {
    "A_row_literal": {"组1": "row_literal", "组2": "row_literal", "组3": "row_literal"},
    "B_full_cube_periodic": {
        "组1": "full_cube_periodic",
        "组2": "full_cube_periodic",
        "组3": "full_cube_periodic",
    },
    "C_thin_12_full_3": {
        "组1": "thin_prism_periodic",
        "组2": "thin_prism_periodic",
        "组3": "full_cube_periodic",
    },
}

INTERNAL_MODES = ("disconnected_fragments", "connected_same_particle")


@dataclass(frozen=True)
class Segment:
    sheet: str
    sheet_row: int
    point1: tuple[float, float, float]
    point2: tuple[float, float, float]
    cylinder: Cylinder


class GraphUnionFind:
    def __init__(self) -> None:
        self.parent: dict[str, str] = {}

    def find(self, node: str) -> str:
        self.parent.setdefault(node, node)
        while self.parent[node] != node:
            self.parent[node] = self.parent[self.parent[node]]
            node = self.parent[node]
        return node

    def union(self, first: str, second: str) -> None:
        first_root = self.find(first)
        second_root = self.find(second)
        if first_root != second_root:
            self.parent[second_root] = first_root


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def project_path(path: Path) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(PROJECT_ROOT.resolve()).as_posix()
    except ValueError:
        return str(resolved)


def load_segments(workbook_path: Path) -> dict[str, list[Segment]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    groups: dict[str, list[Segment]] = {}
    try:
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            segments: list[Segment] = []
            for sheet_row, values in enumerate(
                worksheet.iter_rows(min_row=3, max_col=6, values_only=True), start=3
            ):
                if all(value is None for value in values):
                    continue
                if len(values) != 6 or any(value is None for value in values):
                    raise ValueError(f"{sheet}!{sheet_row} 不是完整六列坐标")
                coordinates = tuple(float(value) for value in values)
                if not all(np.isfinite(coordinates)):
                    raise ValueError(f"{sheet}!{sheet_row} 含非有限坐标")
                point1 = coordinates[:3]
                point2 = coordinates[3:]
                cylinder = Cylinder.from_endpoints(point1, point2, RADIUS_NM)
                segments.append(Segment(sheet, sheet_row, point1, point2, cylinder))
            groups[sheet] = segments
    finally:
        workbook.close()
    return groups


def index_row_map(rows: Iterable[dict[str, Any]]) -> dict[tuple[str, str, int], dict[str, Any]]:
    mapping: dict[tuple[str, str, int], dict[str, Any]] = {}
    for row in rows:
        key = (str(row["sheet"]), str(row["scenario"]), int(row["sheet_row"]))
        if key in mapping:
            raise ValueError(f"重复行映射：{key}")
        mapping[key] = {
            "particle_id": str(row["particle_id"]),
            "component_status": str(row["component_status"]),
            "chain_order": int(row["chain_order"]),
            "orientation_reversed": bool(row["orientation_reversed"]),
            "segment_length_nm": float(row["segment_length_nm"]),
            "component_total_length_nm": float(row["component_total_length_nm"]),
            "component_residual_nm": float(row["component_residual_nm"]),
        }
    return mapping


def index_junctions(rows: Iterable[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    by_scenario: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        parsed = dict(row)
        for field in ("junction_id", "row1", "side1", "row2", "side2", "shift_x", "shift_y", "shift_z"):
            parsed[field] = int(row[field])
        for field in (
            "endpoint1_x",
            "endpoint1_y",
            "endpoint1_z",
            "endpoint2_x",
            "endpoint2_y",
            "endpoint2_z",
            "dot_outward",
        ):
            parsed[field] = float(row[field])
        by_scenario[(str(row["sheet"]), str(row["scenario"]))].append(parsed)
    return by_scenario


def _node_for_row(sheet_row: int) -> str:
    return f"r{sheet_row}"


def _electrode_edge(sheet: str, side: str, segment: Segment, gap_nm: float) -> dict[str, Any]:
    electrode = "LEFT" if side == "left" else "RIGHT"
    node = _node_for_row(segment.sheet_row)
    return {
        "edge_id": f"{sheet}:ELECTRODE:{side.upper()}:r{segment.sheet_row}",
        "sheet": sheet,
        "edge_type": "electrode_contact",
        "node_u": electrode if side == "left" else node,
        "node_v": node if side == "left" else electrode,
        "row_u": "" if side == "left" else segment.sheet_row,
        "row_v": segment.sheet_row if side == "left" else "",
        "electrode": side,
        "gap_lower_nm": gap_nm,
        "gap_upper_nm": gap_nm,
        "gap_estimate_nm": gap_nm,
        "classification": "connected",
        "threshold_relation": "gap<=1.8",
        "method": "closed_form_cylinder_plane",
        "gjk_iterations": 0,
        "gjk_converged": True,
        "capsule_lower_bound_nm": "",
    }


def _orthonormal_plane(axis: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    seed = np.array([1.0, 0.0, 0.0])
    if abs(float(np.dot(seed, axis))) > 0.8:
        seed = np.array([0.0, 1.0, 0.0])
    first = np.cross(axis, seed)
    first /= np.linalg.norm(first)
    return first, np.cross(axis, first)


def slsqp_cylinder_distance(first: Cylinder, second: Cylinder) -> float:
    first_p, first_q = _orthonormal_plane(first.axis)
    second_p, second_q = _orthonormal_plane(second.axis)
    matrix = np.column_stack(
        (
            first.half_length * first.axis,
            first.radius * first_p,
            first.radius * first_q,
            -second.half_length * second.axis,
            -second.radius * second_p,
            -second.radius * second_q,
        )
    )
    center_offset = first.center - second.center
    distance_scale = max(
        1.0,
        float(np.linalg.norm(center_offset)),
        first.characteristic_radius(),
        second.characteristic_radius(),
    )

    def objective(values: np.ndarray) -> float:
        difference = center_offset + matrix @ values
        return float(np.dot(difference, difference)) / distance_scale**2

    def gradient(values: np.ndarray) -> np.ndarray:
        difference = center_offset + matrix @ values
        return 2.0 * matrix.T @ difference / distance_scale**2

    constraints = [
        {
            "type": "ineq",
            "fun": lambda values: 1.0 - values[1] ** 2 - values[2] ** 2,
            "jac": lambda values: np.array(
                [0.0, -2.0 * values[1], -2.0 * values[2], 0.0, 0.0, 0.0]
            ),
        },
        {
            "type": "ineq",
            "fun": lambda values: 1.0 - values[4] ** 2 - values[5] ** 2,
            "jac": lambda values: np.array(
                [0.0, 0.0, 0.0, 0.0, -2.0 * values[4], -2.0 * values[5]]
            ),
        },
    ]
    bounds = [(-1.0, 1.0)] * 6
    unconstrained = np.linalg.lstsq(matrix, -center_offset, rcond=None)[0]

    def make_feasible(values: np.ndarray) -> np.ndarray:
        result = np.clip(np.asarray(values, dtype=float), -1.0, 1.0)
        for first_index, second_index in ((1, 2), (4, 5)):
            radial_norm = float(np.hypot(result[first_index], result[second_index]))
            if radial_norm > 1.0:
                result[[first_index, second_index]] /= radial_norm
        return result

    starts = [np.zeros(6), make_feasible(unconstrained)]
    for first_axial in (-1.0, 1.0):
        for second_axial in (-1.0, 1.0):
            starts.append(
                np.array([first_axial, 0.0, 0.0, second_axial, 0.0, 0.0])
            )

    feasible_results = []
    for start in starts:
        result = minimize(
            objective,
            start,
            jac=gradient,
            bounds=bounds,
            constraints=constraints,
            method="SLSQP",
            options={"ftol": 1e-14, "maxiter": 2000, "disp": False},
        )
        radial_one = result.x[1] ** 2 + result.x[2] ** 2
        radial_two = result.x[4] ** 2 + result.x[5] ** 2
        if radial_one <= 1.0 + 1e-9 and radial_two <= 1.0 + 1e-9:
            feasible_results.append(result)
        if result.success and result.fun <= 1e-24:
            break
    if not feasible_results:
        raise RuntimeError("SLSQP 未返回可行圆柱距离点")
    best = min(feasible_results, key=lambda candidate: candidate.fun)
    return distance_scale * float(np.sqrt(max(best.fun, 0.0)))


def _pair_record(sheet: str, first: Segment, second: Segment) -> dict[str, Any]:
    capsule_gap = capsule_cylinder_distance(first.cylinder, second.cylinder)
    base = {
        "edge_id": f"{sheet}:PAIR:r{first.sheet_row}:r{second.sheet_row}",
        "sheet": sheet,
        "edge_type": "cylinder_contact",
        "node_u": _node_for_row(first.sheet_row),
        "node_v": _node_for_row(second.sheet_row),
        "row_u": first.sheet_row,
        "row_v": second.sheet_row,
        "electrode": "",
        "capsule_lower_bound_nm": capsule_gap,
    }
    if capsule_gap > CONTACT_CUTOFF_NM + SCREEN_GUARD_NM:
        return {
            **base,
            "gap_lower_nm": max(0.0, capsule_gap - SCREEN_GUARD_NM),
            "gap_upper_nm": "",
            "gap_estimate_nm": "",
            "classification": "separated",
            "threshold_relation": "capsule_lower_bound>1.8",
            "method": "capsule_broad_phase_rejection",
            "gjk_iterations": 0,
            "gjk_converged": "",
        }

    bounds = distance_bounds(
        first.cylinder,
        second.cylinder,
        absolute_tolerance=GJK_ABS_TOL_NM,
        relative_tolerance=GJK_REL_TOL,
        max_iterations=512,
    )
    classification = bounds.classify(CONTACT_CUTOFF_NM)
    if classification == "uncertain":
        tighter = distance_bounds(
            first.cylinder,
            second.cylinder,
            absolute_tolerance=1e-12,
            relative_tolerance=1e-14,
            max_iterations=2048,
        )
        if tighter.width < bounds.width or tighter.classify(CONTACT_CUTOFF_NM) != "uncertain":
            bounds = tighter
            classification = bounds.classify(CONTACT_CUTOFF_NM)

    relation = {
        "connected": "gap<=1.8",
        "separated": "gap>1.8",
        "uncertain": "bounds_straddle_1.8",
    }[classification]
    estimate = 0.0 if bounds.upper <= 1e-9 else bounds.estimate
    slsqp_gap = slsqp_cylinder_distance(first.cylinder, second.cylinder)
    slsqp_classification = "connected" if slsqp_gap <= CONTACT_CUTOFF_NM else "separated"
    return {
        **base,
        "gap_lower_nm": bounds.lower,
        "gap_upper_nm": bounds.upper,
        "gap_estimate_nm": estimate,
        "classification": classification,
        "threshold_relation": relation,
        "method": "capsule_then_flat_cylinder_gjk",
        "gjk_iterations": bounds.iterations,
        "gjk_converged": bounds.converged,
        "slsqp_check_gap_nm": slsqp_gap,
        "slsqp_abs_difference_nm": abs(slsqp_gap - estimate),
        "slsqp_same_threshold_class": slsqp_classification == classification,
    }


def compute_physical_geometry(sheet: str, segments: list[Segment]) -> dict[str, Any]:
    contacts: list[dict[str, Any]] = []
    narrow_phase: list[dict[str, Any]] = []
    broad_rejected = 0

    for segment in segments:
        left_gap = shape_plane_distance(segment.cylinder, np.array([1.0, 0.0, 0.0]), -ELECTRODE_X_NM)
        right_gap = shape_plane_distance(segment.cylinder, np.array([1.0, 0.0, 0.0]), ELECTRODE_X_NM)
        if left_gap <= CONTACT_CUTOFF_NM:
            contacts.append(_electrode_edge(sheet, "left", segment, left_gap))
        if right_gap <= CONTACT_CUTOFF_NM:
            contacts.append(_electrode_edge(sheet, "right", segment, right_gap))

    for first_index, first in enumerate(segments):
        for second in segments[first_index + 1 :]:
            record = _pair_record(sheet, first, second)
            if record["method"] == "capsule_broad_phase_rejection":
                broad_rejected += 1
                continue
            narrow_phase.append(record)
            if record["classification"] == "connected":
                contacts.append(record)

    uncertain = [row for row in narrow_phase if row["classification"] == "uncertain"]
    possible_contacts = contacts + uncertain
    class_counts: dict[str, int] = defaultdict(int)
    for row in narrow_phase:
        class_counts[row["classification"]] += 1
    return {
        "contacts": contacts,
        "possible_contacts": possible_contacts,
        "narrow_phase": narrow_phase,
        "screening": {
            "sheet": sheet,
            "segment_count": len(segments),
            "total_pairs": len(segments) * (len(segments) - 1) // 2,
            "broad_phase_rejected": broad_rejected,
            "narrow_phase_pairs": len(narrow_phase),
            "narrow_connected": class_counts["connected"],
            "narrow_separated": class_counts["separated"],
            "narrow_uncertain": class_counts["uncertain"],
            "gjk_nonconverged_narrow": sum(not bool(row["gjk_converged"]) for row in narrow_phase),
            "slsqp_threshold_disagreements": sum(
                row["slsqp_same_threshold_class"] is False for row in narrow_phase
            ),
            "slsqp_max_abs_difference_nm": max(
                (float(row["slsqp_abs_difference_nm"]) for row in narrow_phase), default=None
            ),
            "max_narrow_bound_width_nm": max(
                (float(row["gap_upper_nm"]) - float(row["gap_lower_nm"]) for row in narrow_phase),
                default=None,
            ),
            "min_threshold_decision_margin_nm": min(
                (
                    CONTACT_CUTOFF_NM - float(row["gap_upper_nm"])
                    if row["classification"] == "connected"
                    else float(row["gap_lower_nm"]) - CONTACT_CUTOFF_NM
                    for row in narrow_phase
                    if row["classification"] != "uncertain"
                ),
                default=None,
            ),
            "electrode_contacts": sum(row["edge_type"] == "electrode_contact" for row in contacts),
            "physical_contact_edges": sum(row["edge_type"] == "cylinder_contact" for row in contacts),
            "max_connected_gap_upper_nm": max(
                (float(row["gap_upper_nm"]) for row in contacts if row["edge_type"] == "cylinder_contact"),
                default=None,
            ),
            "min_narrow_separated_gap_lower_nm": min(
                (float(row["gap_lower_nm"]) for row in narrow_phase if row["classification"] == "separated"),
                default=None,
            ),
        },
    }


def _particle_lookup(
    sheet: str,
    reconstruction_scenario: str,
    segments: list[Segment],
    row_map: dict[tuple[str, str, int], dict[str, Any]],
) -> dict[int, dict[str, Any]]:
    result: dict[int, dict[str, Any]] = {}
    for segment in segments:
        key = (sheet, reconstruction_scenario, segment.sheet_row)
        if key not in row_map:
            raise KeyError(f"缺少行身份映射：{key}")
        result[segment.sheet_row] = row_map[key]
    return result


def _periodic_lengths(reconstruction_scenario: str) -> np.ndarray:
    if reconstruction_scenario == "thin_prism_periodic":
        return np.array([10000.0, 1000.0, 1000.0])
    return np.array([10000.0, 10000.0, 10000.0])


def build_internal_edges(
    sheet: str,
    reconstruction_scenario: str,
    particle_lookup: dict[int, dict[str, Any]],
    junctions_by_scenario: dict[tuple[str, str], list[dict[str, Any]]],
    enabled: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    graph_edges: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    if reconstruction_scenario == "row_literal":
        return graph_edges, audit_rows

    lengths = _periodic_lengths(reconstruction_scenario)
    for junction in junctions_by_scenario.get((sheet, reconstruction_scenario), []):
        row1 = junction["row1"]
        row2 = junction["row2"]
        particle1 = particle_lookup[row1]
        particle2 = particle_lookup[row2]
        if particle1["particle_id"] != particle2["particle_id"]:
            raise ValueError(f"接头跨越两个 particle_id：{sheet} r{row1}, r{row2}")
        endpoint1 = np.array(
            [junction["endpoint1_x"], junction["endpoint1_y"], junction["endpoint1_z"]]
        )
        endpoint2 = np.array(
            [junction["endpoint2_x"], junction["endpoint2_y"], junction["endpoint2_z"]]
        )
        shift = np.array([junction["shift_x"], junction["shift_y"], junction["shift_z"]])
        residual = float(np.linalg.norm(endpoint1 + shift * lengths - endpoint2))
        edge = {
            "edge_id": f"{sheet}:INTERNAL:{reconstruction_scenario}:j{junction['junction_id']}",
            "sheet": sheet,
            "edge_type": "same_particle_periodic_junction",
            "node_u": _node_for_row(row1),
            "node_v": _node_for_row(row2),
            "row_u": row1,
            "row_v": row2,
            "electrode": "",
            "gap_lower_nm": "",
            "gap_upper_nm": "",
            "gap_estimate_nm": "",
            "classification": "internal_assumption" if enabled else "not_in_graph",
            "threshold_relation": "not_applicable_same_particle",
            "method": "unique_periodic_endpoint_mapping",
            "gjk_iterations": "",
            "gjk_converged": "",
            "capsule_lower_bound_nm": "",
            "particle_id": particle1["particle_id"],
            "component_status": particle1["component_status"],
            "crossing_axes": junction["crossing_axes"],
            "shift_x": junction["shift_x"],
            "shift_y": junction["shift_y"],
            "shift_z": junction["shift_z"],
            "mapped_endpoint_residual_nm": residual,
            "enabled": enabled,
        }
        audit_rows.append(edge)
        if enabled:
            graph_edges.append(edge)
    return graph_edges, audit_rows


def annotate_physical_edge(edge: dict[str, Any], lookup: dict[int, dict[str, Any]]) -> dict[str, Any]:
    result = dict(edge)
    row_u = edge["row_u"]
    row_v = edge["row_v"]
    result["particle_u"] = lookup[int(row_u)]["particle_id"] if row_u != "" else edge["node_u"]
    result["particle_v"] = lookup[int(row_v)]["particle_id"] if row_v != "" else edge["node_v"]
    return result


def connected_by_union_find(edges: Iterable[dict[str, Any]]) -> bool:
    union_find = GraphUnionFind()
    union_find.find("LEFT")
    union_find.find("RIGHT")
    for edge in edges:
        union_find.union(str(edge["node_u"]), str(edge["node_v"]))
    return union_find.find("LEFT") == union_find.find("RIGHT")


def shortest_path(edges: Iterable[dict[str, Any]]) -> tuple[list[str] | None, list[dict[str, Any]]]:
    adjacency: dict[str, list[tuple[str, dict[str, Any]]]] = defaultdict(list)
    for edge in edges:
        first = str(edge["node_u"])
        second = str(edge["node_v"])
        adjacency[first].append((second, edge))
        adjacency[second].append((first, edge))
    for values in adjacency.values():
        values.sort(key=lambda item: (item[0], item[1]["edge_id"]))

    queue: list[tuple[int, int, tuple[str, ...], str]] = [(0, 0, ("LEFT",), "LEFT")]
    best_cost: dict[str, tuple[int, int, tuple[str, ...]]] = {"LEFT": (0, 0, ("LEFT",))}
    predecessor: dict[str, tuple[str, dict[str, Any]] | None] = {"LEFT": None}
    while queue:
        residual_edges, edge_count, signature, current = heapq.heappop(queue)
        if best_cost.get(current) != (residual_edges, edge_count, signature):
            continue
        if current == "RIGHT":
            break
        for neighbor, edge in adjacency.get(current, []):
            residual_increment = int(
                edge.get("edge_type") == "same_particle_periodic_junction"
                and edge.get("component_status") != "unique"
            )
            candidate = (
                residual_edges + residual_increment,
                edge_count + 1,
                signature + (neighbor,),
            )
            if neighbor in best_cost and best_cost[neighbor] <= candidate:
                continue
            best_cost[neighbor] = candidate
            predecessor[neighbor] = (current, edge)
            heapq.heappush(queue, (*candidate, neighbor))
    if "RIGHT" not in predecessor:
        return None, []

    nodes = ["RIGHT"]
    path_edges: list[dict[str, Any]] = []
    while nodes[-1] != "LEFT":
        previous, edge = predecessor[nodes[-1]]
        path_edges.append(edge)
        nodes.append(previous)
    nodes.reverse()
    path_edges.reverse()
    return nodes, path_edges


def particle_path(nodes: list[str] | None, lookup: dict[int, dict[str, Any]]) -> list[dict[str, Any]] | None:
    if nodes is None:
        return None
    compressed: list[dict[str, Any]] = []
    for node in nodes:
        if node in {"LEFT", "RIGHT"}:
            compressed.append({"node": node})
            continue
        row_number = int(node[1:])
        particle_id = lookup[row_number]["particle_id"]
        if compressed and compressed[-1].get("particle_id") == particle_id:
            compressed[-1]["rows"].append(row_number)
        else:
            compressed.append(
                {
                    "particle_id": particle_id,
                    "component_status": lookup[row_number]["component_status"],
                    "rows": [row_number],
                }
            )
    return compressed


def evaluate_scenario(
    scenario: str,
    sheet: str,
    reconstruction_scenario: str,
    internal_mode: str,
    segments: list[Segment],
    geometry: dict[str, Any],
    row_map: dict[tuple[str, str, int], dict[str, Any]],
    junctions_by_scenario: dict[tuple[str, str], list[dict[str, Any]]],
) -> dict[str, Any]:
    lookup = _particle_lookup(sheet, reconstruction_scenario, segments, row_map)
    internal_enabled = internal_mode == "connected_same_particle"
    internal_edges, internal_audit = build_internal_edges(
        sheet,
        reconstruction_scenario,
        lookup,
        junctions_by_scenario,
        internal_enabled,
    )
    definite_physical = [annotate_physical_edge(edge, lookup) for edge in geometry["contacts"]]
    possible_physical = [annotate_physical_edge(edge, lookup) for edge in geometry["possible_contacts"]]
    definite_edges = definite_physical + internal_edges
    possible_edges = possible_physical + internal_edges
    definite_connected = connected_by_union_find(definite_edges)
    possible_connected = connected_by_union_find(possible_edges)
    definite_nodes, definite_path_edges = shortest_path(definite_edges)
    possible_nodes, possible_path_edges = shortest_path(possible_edges)
    if definite_connected != (definite_nodes is not None):
        raise AssertionError("并查集与确定边见证路径判定不一致")
    if possible_connected != (possible_nodes is not None):
        raise AssertionError("并查集与可能边见证路径判定不一致")

    if definite_connected:
        conclusion = "conductive"
        witness_nodes = definite_nodes
        witness_edges = definite_path_edges
    elif possible_nodes is None:
        conclusion = "nonconductive"
        witness_nodes = None
        witness_edges = []
    else:
        conclusion = "numerically_undetermined"
        witness_nodes = possible_nodes
        witness_edges = possible_path_edges

    particle_ids = {value["particle_id"] for value in lookup.values()}
    residual_particles = {
        value["particle_id"]
        for value in lookup.values()
        if value["component_status"] == "cannot_reconstruct"
    }
    return {
        "scenario": scenario,
        "sheet": sheet,
        "reconstruction_scenario": reconstruction_scenario,
        "internal_mode": internal_mode,
        "row_count": len(segments),
        "particle_count": len(particle_ids),
        "residual_particle_count": len(residual_particles),
        "periodic_junction_count": len(internal_audit),
        "internal_edges_enabled": len(internal_edges),
        "physical_edges": len(definite_physical),
        "uncertain_physical_edges": len(possible_physical) - len(definite_physical),
        "conductive_definite": definite_connected,
        "conductive_possible": possible_connected,
        "conclusion": conclusion,
        "witness_row_nodes": witness_nodes,
        "witness_particle_path": particle_path(witness_nodes, lookup),
        "witness_edges": witness_edges,
        "all_graph_edges": definite_edges,
        "internal_junction_audit": internal_audit,
        "row_lookup": lookup,
    }


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _json_ready(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_ready(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_ready(item) for item in value]
    if isinstance(value, np.generic):
        return value.item()
    return value


def _format_path(nodes: list[str] | None) -> str:
    if nodes is None:
        return "-"
    return " -> ".join(nodes)


def _format_particle_path(path: list[dict[str, Any]] | None) -> str:
    if path is None:
        return "-"
    parts: list[str] = []
    for item in path:
        if "node" in item:
            parts.append(item["node"])
        else:
            rows = ",".join(f"r{row}" for row in item["rows"])
            parts.append(f"{item['particle_id']}[{rows};{item['component_status']}]")
    return " -> ".join(parts)


def build_report(
    metadata: dict[str, Any],
    screening_rows: list[dict[str, Any]],
    evaluations: list[dict[str, Any]],
) -> str:
    lines = [
        "# Q1 正式逐行独立介质导通判定报告",
        "",
        "> **正式口径：仅 `A_row_literal / disconnected_fragments`。** "
        "附件中的每行及其截断片段均作为独立介质；"
        "`connected_same_particle` 仅用于程序一致性检查，不进入本题结果注册表。",
        "",
        "## 计算口径",
        "",
        "- 每条 Excel 记录按给定端点直接构造半径 30 nm 的有限平底圆柱；短段不拉长、不补段。",
        "- 圆柱对先用胶囊距离作下界宽相；只有下界不排除 1.8 nm 接触时才调用平底圆柱 GJK 窄相。",
        "- 只使用基础域内记录的直接欧氏距离；没有对任意两行使用全局 minimum-image。",
        "- 电极 x=±5000 nm 采用圆柱投影区间的闭式距离。",
        "- 每条记录及其截断片段均作为独立介质，基础域内只按实际实体位置判定接触。",
        "- 截断后不保留同源内部连接；接触图仅加入实体间距或实体到电极距离不超过阈值的边。",
        "",
        "## 几何筛选审计",
        "",
        "|组|记录数|总对数|胶囊排除|GJK窄相|接触|窄相分离|不确定|电极边|",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in screening_rows:
        lines.append(
            f"|{row['sheet']}|{row['segment_count']}|{row['total_pairs']}|"
            f"{row['broad_phase_rejected']}|{row['narrow_phase_pairs']}|"
            f"{row['narrow_connected']}|{row['narrow_separated']}|"
            f"{row['narrow_uncertain']}|{row['electrode_contacts']}|"
        )

    lines.extend(
        [
            "",
            "全部窄相对另用六变量凸约束 SLSQP 独立复核；该检查不替代 GJK 上下界证书。"
            f"三组阈值分类不一致总数为 {sum(row['slsqp_threshold_disagreements'] for row in screening_rows)}，"
            f"最大距离估计差为 {max(row['slsqp_max_abs_difference_nm'] or 0.0 for row in screening_rows):.3e} nm。",
            f"GJK 中共有 {sum(row['gjk_nonconverged_narrow'] for row in screening_rows)} 对未触发收敛标志；"
            f"所有窄相上下界最大宽度为 {max(row['max_narrow_bound_width_nm'] or 0.0 for row in screening_rows):.3e} nm，"
            f"距 1.8 nm 阈值的最小确定裕量为 {min(row['min_threshold_decision_margin_nm'] for row in screening_rows if row['min_threshold_decision_margin_nm'] is not None):.6f} nm，"
            "故没有分类不确定边。",
            "",
            "## 导通结果",
            "",
            "|解释|内部连接|组|行数|particle数|接头数|结论|行号见证路径|",
            "|---|---|---|---:|---:|---:|---|---|",
        ]
    )
    for result in evaluations:
        lines.append(
            f"|{result['scenario']}|{result['internal_mode']}|{result['sheet']}|"
            f"{result['row_count']}|{result['particle_count']}|{result['periodic_junction_count']}|"
            f"{result['conclusion']}|{_format_path(result['witness_row_nodes'])}|"
        )

    lines.extend(["", "## 粒子/行号见证路径", ""])
    for result in evaluations:
        lines.append(
            f"- `{result['scenario']} / {result['internal_mode']} / {result['sheet']}`："
            f"{_format_particle_path(result['witness_particle_path'])}"
        )

    lines.extend(
        [
            "",
            "## 边证据说明",
            "",
            "- `physical_contact_edges.csv` 给出全部确定物理接触边的最小间隙上下界和 1.8 nm 关系。",
            "- `narrow_phase_pairs.csv` 还保留胶囊未能排除、经 GJK 判为分离或不确定的全部近邻对。",
            "- `independent_slsqp_validation.csv` 给出全部窄相对的独立凸优化距离检查及阈值分类一致性。",
            "- `scenario_graph_edges.csv` 给出每种解释/内部模式实际使用的全部图边及两端 particle_id。",
            "- `witness_path_edges.csv` 是每条导通见证路径的逐边证据。周期内部边不以 1.8 nm 判定，间隙列为空，另报映射后端点残差。",
            "- `internal_junction_audit.csv` 同时列出启用与禁用模式下每个接头，便于核对未闭合链没有被补段。",
            "",
            "## 可复现性",
            "",
            f"- Python：`{metadata['python_version']}`；NumPy：`{metadata['numpy_version']}`；openpyxl：`{metadata['openpyxl_version']}`。",
            f"- 附件 SHA-256：`{metadata['input_hashes']['workbook']}`。",
            f"- 几何核 SHA-256：`{metadata['input_hashes']['geometry_kernel']}`。",
            f"- 运行耗时：`{metadata['elapsed_seconds']:.3f} s`。",
            "",
            "```powershell",
            "# 从项目根目录运行以下命令",
            "python -m unittest discover -s tests -v",
            "python 问题\\问题1\\src\\solve.py",
            "```",
            "",
        ]
    )
    return "\n".join(lines)


def run(args: argparse.Namespace) -> dict[str, Any]:
    start = time.perf_counter()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    reconstruction = reconstruct_workbook(args.xlsx)
    write_processed_outputs(reconstruction, args.processed_dir)
    segments_by_sheet = load_segments(args.xlsx)
    row_map = index_row_map(reconstruction["row_identity_map"])
    junctions = index_junctions(reconstruction["junctions"])

    physical_by_sheet: dict[str, dict[str, Any]] = {}
    for sheet, segments in segments_by_sheet.items():
        physical_by_sheet[sheet] = compute_physical_geometry(sheet, segments)

    evaluations: list[dict[str, Any]] = []
    for scenario, mapping_by_sheet in SCENARIOS.items():
        for internal_mode in INTERNAL_MODES:
            for sheet, segments in segments_by_sheet.items():
                evaluations.append(
                    evaluate_scenario(
                        scenario,
                        sheet,
                        mapping_by_sheet[sheet],
                        internal_mode,
                        segments,
                        physical_by_sheet[sheet],
                        row_map,
                        junctions,
                    )
                )

    screening_rows = [physical_by_sheet[sheet]["screening"] for sheet in segments_by_sheet]
    contact_rows = [row for sheet in segments_by_sheet for row in physical_by_sheet[sheet]["contacts"]]
    narrow_rows = [row for sheet in segments_by_sheet for row in physical_by_sheet[sheet]["narrow_phase"]]

    summary_rows: list[dict[str, Any]] = []
    scenario_edge_rows: list[dict[str, Any]] = []
    witness_rows: list[dict[str, Any]] = []
    internal_rows: list[dict[str, Any]] = []
    assignment_rows: list[dict[str, Any]] = []
    for result in evaluations:
        summary_rows.append(
            {
                key: result[key]
                for key in (
                    "scenario",
                    "sheet",
                    "reconstruction_scenario",
                    "internal_mode",
                    "row_count",
                    "particle_count",
                    "residual_particle_count",
                    "periodic_junction_count",
                    "internal_edges_enabled",
                    "physical_edges",
                    "uncertain_physical_edges",
                    "conductive_definite",
                    "conductive_possible",
                    "conclusion",
                )
            }
            | {
                "witness_row_nodes": _format_path(result["witness_row_nodes"]),
                "witness_particle_path": _format_particle_path(result["witness_particle_path"]),
            }
        )
        for edge in result["all_graph_edges"]:
            scenario_edge_rows.append(
                {
                    "scenario": result["scenario"],
                    "reconstruction_scenario": result["reconstruction_scenario"],
                    "internal_mode": result["internal_mode"],
                    **edge,
                }
            )
        for edge_index, edge in enumerate(result["witness_edges"], start=1):
            witness_rows.append(
                {
                    "scenario": result["scenario"],
                    "reconstruction_scenario": result["reconstruction_scenario"],
                    "internal_mode": result["internal_mode"],
                    "path_edge_order": edge_index,
                    **edge,
                }
            )
        for edge in result["internal_junction_audit"]:
            internal_rows.append(
                {
                    "scenario": result["scenario"],
                    "reconstruction_scenario": result["reconstruction_scenario"],
                    "internal_mode": result["internal_mode"],
                    **edge,
                }
            )
        for sheet_row, mapping in sorted(result["row_lookup"].items()):
            assignment_rows.append(
                {
                    "scenario": result["scenario"],
                    "reconstruction_scenario": result["reconstruction_scenario"],
                    "internal_mode": result["internal_mode"],
                    "sheet": result["sheet"],
                    "sheet_row": sheet_row,
                    **mapping,
                }
            )

    edge_fields = [
        "edge_id",
        "sheet",
        "edge_type",
        "node_u",
        "node_v",
        "row_u",
        "row_v",
        "electrode",
        "particle_u",
        "particle_v",
        "particle_id",
        "component_status",
        "gap_lower_nm",
        "gap_upper_nm",
        "gap_estimate_nm",
        "classification",
        "threshold_relation",
        "method",
        "capsule_lower_bound_nm",
        "gjk_iterations",
        "gjk_converged",
        "slsqp_check_gap_nm",
        "slsqp_abs_difference_nm",
        "slsqp_same_threshold_class",
        "crossing_axes",
        "shift_x",
        "shift_y",
        "shift_z",
        "mapped_endpoint_residual_nm",
        "enabled",
    ]
    write_csv(output_dir / "screening_summary.csv", screening_rows, list(screening_rows[0]))
    write_csv(output_dir / "physical_contact_edges.csv", contact_rows, edge_fields)
    write_csv(output_dir / "narrow_phase_pairs.csv", narrow_rows, edge_fields)
    write_csv(
        output_dir / "independent_slsqp_validation.csv",
        narrow_rows,
        [
            "sheet",
            "row_u",
            "row_v",
            "classification",
            "gap_lower_nm",
            "gap_upper_nm",
            "gap_estimate_nm",
            "gjk_converged",
            "slsqp_check_gap_nm",
            "slsqp_abs_difference_nm",
            "slsqp_same_threshold_class",
        ],
    )
    write_csv(output_dir / "scenario_summary.csv", summary_rows, list(summary_rows[0]))
    write_csv(
        output_dir / "scenario_graph_edges.csv",
        scenario_edge_rows,
        ["scenario", "reconstruction_scenario", "internal_mode"] + edge_fields,
    )
    write_csv(
        output_dir / "witness_path_edges.csv",
        witness_rows,
        ["scenario", "reconstruction_scenario", "internal_mode", "path_edge_order"] + edge_fields,
    )
    write_csv(
        output_dir / "internal_junction_audit.csv",
        internal_rows,
        ["scenario", "reconstruction_scenario", "internal_mode"] + edge_fields,
    )
    write_csv(
        output_dir / "row_assignments.csv",
        assignment_rows,
        [
            "scenario",
            "reconstruction_scenario",
            "internal_mode",
            "sheet",
            "sheet_row",
            "particle_id",
            "component_status",
            "chain_order",
            "orientation_reversed",
            "segment_length_nm",
            "component_total_length_nm",
            "component_residual_nm",
        ],
    )

    elapsed = time.perf_counter() - start
    metadata = {
        "generated_at_local": time.strftime("%Y-%m-%d %H:%M:%S %z"),
        "elapsed_seconds": elapsed,
        "python_version": platform.python_version(),
        "numpy_version": np.__version__,
        "openpyxl_version": openpyxl.__version__,
        "constants": {
            "radius_nm": RADIUS_NM,
            "contact_cutoff_nm": CONTACT_CUTOFF_NM,
            "electrode_x_nm": [-ELECTRODE_X_NM, ELECTRODE_X_NM],
            "gjk_absolute_tolerance_nm": GJK_ABS_TOL_NM,
            "gjk_relative_tolerance": GJK_REL_TOL,
            "capsule_screen_guard_nm": SCREEN_GUARD_NM,
        },
        "input_paths": {
            "workbook": project_path(args.xlsx),
            "processed_dir": project_path(args.processed_dir),
            "reconstruction_source": project_path(RECONSTRUCTION_DIR / "reconstruct_segments.py"),
            "geometry_kernel": project_path(COMMON_DIR / "geometry_kernel.py"),
        },
        "input_hashes": {
            "workbook": sha256(args.xlsx),
            "row_map": sha256(args.processed_dir / "row_identity_map.csv"),
            "junctions": sha256(args.processed_dir / "junctions.csv"),
            "reconstruction_source": sha256(RECONSTRUCTION_DIR / "reconstruct_segments.py"),
            "geometry_kernel": sha256(COMMON_DIR / "geometry_kernel.py"),
        },
        "scenario_definitions": SCENARIOS,
        "internal_mode_definitions": {
            "disconnected_fragments": "不添加周期接头内部边",
            "connected_same_particle": "仅添加 junctions.csv 中一对一共线已选接头；不补缺失片段",
        },
    }
    result_payload = {
        "metadata": metadata,
        "screening": screening_rows,
        "scenario_results": [
            {key: value for key, value in result.items() if key not in {"all_graph_edges", "internal_junction_audit", "row_lookup"}}
            for result in evaluations
        ],
    }
    (output_dir / "q1_results.json").write_text(
        json.dumps(_json_ready(result_payload), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / "report.md").write_text(
        build_report(metadata, screening_rows, evaluations), encoding="utf-8"
    )
    if not args.skip_registry:
        register_q1_results(result_payload)
    return result_payload


def register_q1_results(
    payload: dict[str, Any],
    *,
    registry_path: Path = DEFAULT_REGISTRY,
    latex_path: Path = DEFAULT_TEX,
) -> None:
    indexed = {
        (row["scenario"], row["internal_mode"], row["sheet"]): row
        for row in payload["scenario_results"]
    }
    artifact = "问题/问题1/results/q1_results.json"
    validation = (
        "每行及截断片段按独立介质处理；tests/test_q1_solver.py；"
        "GJK距离界与全部窄相SLSQP独立阈值复核"
    )
    source = "问题/问题1/src/solve.py"

    registry = read_registry(registry_path)
    registry["results"] = {
        key: value
        for key, value in registry["results"].items()
        if not key.startswith("q1_")
    }
    write_registry(registry, registry_path)

    for group_index, sheet in enumerate(("组1", "组2", "组3"), start=1):
        result = indexed[("A_row_literal", "disconnected_fragments", sheet)]
        conductive = bool(result["conductive_definite"])
        register_result(
            f"q1_fragment_disconnected_group{group_index}_conductive",
            question=1,
            value=conductive,
            formatted="导通" if conductive else "不导通",
            source_script=source,
            source_artifact=artifact,
            validation=validation,
            registry_path=registry_path,
        )
        if conductive:
            register_result(
                f"q1_fragment_disconnected_group{group_index}_witness",
                question=1,
                value=_format_path(result["witness_row_nodes"]),
                source_script=source,
                source_artifact="问题/问题1/results/witness_path_edges.csv",
                validation=validation,
                registry_path=registry_path,
            )

    screening = payload["screening"]
    register_result(
        "q1_numerically_uncertain_edges",
        question=1,
        value=sum(int(row["narrow_uncertain"]) for row in screening),
        unit="条",
        source_script=source,
        source_artifact="问题/问题1/results/screening_summary.csv",
        validation=validation,
        registry_path=registry_path,
    )
    register_result(
        "q1_slsqp_threshold_disagreements",
        question=1,
        value=sum(int(row["slsqp_threshold_disagreements"]) for row in screening),
        unit="条",
        source_script=source,
        source_artifact="问题/问题1/results/independent_slsqp_validation.csv",
        validation=validation,
        registry_path=registry_path,
    )
    closest_gap = max(
        float(row["max_connected_gap_upper_nm"]) for row in screening
    )
    register_result(
        "q1_closest_connected_gap_upper_nm",
        question=1,
        value=closest_gap,
        formatted=f"{closest_gap:.10f}",
        unit="nm",
        source_script=source,
        source_artifact="问题/问题1/results/screening_summary.csv",
        validation=validation,
        registry_path=registry_path,
    )
    export_latex(registry_path, latex_path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Q1 逐行独立介质的有限平底圆柱导通判定"
    )
    parser.add_argument("--xlsx", type=Path, default=PROJECT_ROOT / "00_赛题与附件" / "附件.xlsx")
    parser.add_argument(
        "--processed-dir", type=Path, default=PROJECT_ROOT / "02_数据与参数" / "processed"
    )
    parser.add_argument("--output-dir", type=Path, default=QUESTION_ROOT / "results")
    parser.add_argument("--skip-registry", action="store_true")
    return parser.parse_args(argv)


if __name__ == "__main__":
    run(parse_args())
